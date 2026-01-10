#!/usr/bin/env python3
"""
Convert family data from Excel spreadsheet to GEDCOM format.

Expected Excel format:
- Sheet: "People" - Individual records
  Columns: ID, Name, Sex, Birth Date, Birth Place, Death Date, Death Place
  
- Sheet: "Families" - Family relationships  
  Columns: Family ID, Husband ID, Wife ID, Children IDs (comma-separated)
  
- Sheet: "Marriages" - Marriage information
  Columns: Family ID, Marriage Date, Marriage Place
"""

import openpyxl
import sys
from datetime import datetime
from pathlib import Path


class GEDCOMGenerator:
    def __init__(self):
        self.ged_lines = []
        self.person_count = 0
        self.family_count = 0
        
    def add_line(self, level, tag, value="", xref=""):
        """Add a GEDCOM line with proper formatting"""
        if xref:
            line = f"{level} @{xref}@ {tag}"
        else:
            line = f"{level} {tag}"
        if value:
            # Escape special characters in values
            value_str = str(value).replace("\n", " ")
            line += f" {value_str}"
        self.ged_lines.append(line)
    
    def create_header(self):
        """Create GEDCOM header"""
        self.add_line(0, "HEAD")
        self.add_line(1, "SOUR", "FamilyTreeApp")
        self.add_line(1, "DEST", "Gramps")
        self.add_line(1, "DATE", datetime.now().strftime("%d %b %Y"))
        self.add_line(2, "TIME", datetime.now().strftime("%H:%M:%S"))
        self.add_line(1, "VERS", "5.3.1")
        self.add_line(1, "CHAR", "UTF-8")
        self.add_line(1, "LANG", "English")
        self.add_line(0, "TRLR")
    
    def add_person(self, person_id, name, sex="", birth_date="", birth_place="", 
                   death_date="", death_place=""):
        """Add a person (INDI) record"""
        self.add_line(0, "INDI", xref=person_id)
        
        # Name format: First /Surname/
        if name:
            parts = str(name).strip().split()
            if len(parts) > 1:
                first = " ".join(parts[:-1])
                last = parts[-1]
                self.add_line(1, "NAME", f"{first} /{last}/")
            else:
                self.add_line(1, "NAME", f"{name} //")
        else:
            self.add_line(1, "NAME", " //")
        
        if sex and sex.upper() in ("M", "F"):
            self.add_line(1, "SEX", sex.upper())
        
        if birth_date or birth_place:
            self.add_line(1, "BIRT")
            if birth_date:
                self.add_line(2, "DATE", self._format_date(birth_date))
            if birth_place:
                self.add_line(2, "PLAC", birth_place)
        
        if death_date or death_place:
            self.add_line(1, "DEAT")
            if death_date:
                self.add_line(2, "DATE", self._format_date(death_date))
            if death_place:
                self.add_line(2, "PLAC", death_place)
    
    def add_family(self, family_id, husband_id="", wife_id="", children_ids=None,
                   marriage_date="", marriage_place=""):
        """Add a family (FAM) record"""
        if children_ids is None:
            children_ids = []
        
        self.add_line(0, "FAM", xref=family_id)
        
        if husband_id:
            self.add_line(1, "HUSB", xref=husband_id)
        
        if wife_id:
            self.add_line(1, "WIFE", xref=wife_id)
        
        if marriage_date or marriage_place:
            self.add_line(1, "MARR")
            if marriage_date:
                self.add_line(2, "DATE", self._format_date(marriage_date))
            if marriage_place:
                self.add_line(2, "PLAC", marriage_place)
        
        for child_id in children_ids:
            if child_id and str(child_id).strip():
                self.add_line(1, "CHIL", xref=str(child_id).strip())
    
    def _format_date(self, date_str):
        """Convert various date formats to GEDCOM format (DD MMM YYYY)"""
        if not date_str or str(date_str).strip() == "" or str(date_str).lower() == "nan":
            return ""
        
        date_str = str(date_str).strip()
        
        # Try parsing common formats
        formats = [
            "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
            "%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d",
            "%d/%m/%y", "%d-%m-%y",
            "%B %d, %Y", "%b %d, %Y",
            "%d %B %Y", "%d %b %Y",
        ]
        
        for fmt in formats:
            try:
                dt = datetime.strptime(date_str, fmt)
                return dt.strftime("%d %b %Y")
            except ValueError:
                continue
        
        # If no format matches, return as-is
        return date_str
    
    def get_ged_content(self):
        """Return complete GEDCOM content"""
        return "\n".join(self.ged_lines)
    
    def save(self, filepath):
        """Save GEDCOM to file"""
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(self.get_ged_content())
        print(f"[OK] GEDCOM file saved to: {filepath}")


def convert_excel_to_ged(excel_path, output_ged_path):
    """Main conversion function"""
    
    print(f"Reading Excel file: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        print(f"[ERROR] File not found: {excel_path}")
        return False
    except Exception as e:
        print(f"[ERROR] Could not read Excel file: {e}")
        return False
    
    gen = GEDCOMGenerator()
    person_id_map = {}  # Maps original IDs to I1, I2, etc.
    
    # Process Individuals sheet (or People)
    individuals_sheet = None
    if "Individuals" in wb.sheetnames:
        individuals_sheet = "Individuals"
    elif "People" in wb.sheetnames:
        individuals_sheet = "People"
    
    if individuals_sheet:
        print(f"Processing {individuals_sheet} sheet...")
        ws = wb[individuals_sheet]
        
        # Get headers
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value.lower()] = col_idx
        
        # Process rows
        person_counter = 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row[0]:  # Skip if ID is empty
                continue
            
            original_id = str(row[0]).strip()
            try:
                numeric_id = int(float(original_id))
            except (ValueError, TypeError):
                numeric_id = person_counter
            
            # Create standardized ID
            ged_person_id = f"I{numeric_id}"
            person_id_map[original_id] = ged_person_id
            
            # Extract fields (handle case-insensitive headers)
            name = ""
            sex = ""
            birth_date = ""
            birth_place = ""
            death_date = ""
            death_place = ""
            
            for key in headers:
                idx = headers[key] - 1
                val = row[idx] if idx < len(row) else None
                
                if "name" in key and not name:
                    name = val
                elif "sex" in key and not sex:
                    sex = val
                elif "birth date" in key and not birth_date:
                    birth_date = val
                elif "birth place" in key and not birth_place:
                    birth_place = val
                elif "death date" in key and not death_date:
                    death_date = val
                elif "death place" in key and not death_place:
                    death_place = val
            
            gen.add_person(ged_person_id, name, sex, birth_date, birth_place, death_date, death_place)
            print(f"  Added person: {ged_person_id} - {name}")
            person_counter += 1
    else:
        print("[WARNING] 'Individuals' or 'People' sheet not found")
    
    # Process Families sheet
    if "Families" in wb.sheetnames:
        print("Processing Families sheet...")
        ws = wb["Families"]
        
        # Get headers
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value.lower()] = col_idx
        
        # Process rows
        fam_counter = 1
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row[0]:  # Skip if Family ID is empty
                continue
            
            family_id = str(row[0]).strip()
            husband_id = ""
            wife_id = ""
            children_str = ""
            marriage_date = ""
            marriage_place = ""
            
            for key in headers:
                idx = headers[key] - 1
                val = row[idx] if idx < len(row) else None
                
                if "husband id" in key and not husband_id:
                    husband_id = val
                elif "wife id" in key and not wife_id:
                    wife_id = val
                elif "children" in key and not children_str:
                    children_str = val
                elif "marriage date" in key and not marriage_date:
                    marriage_date = val
                elif "marriage place" in key and not marriage_place:
                    marriage_place = val
            
            # Create standardized family ID
            ged_family_id = f"F{fam_counter}"
            
            # Map person IDs to GED IDs
            if husband_id:
                original_husband = str(husband_id).strip()
                if original_husband in person_id_map:
                    husband_id = person_id_map[original_husband]
                else:
                    try:
                        husband_id = f"I{int(float(original_husband))}"
                    except (ValueError, TypeError):
                        husband_id = ""
            
            if wife_id:
                original_wife = str(wife_id).strip()
                if original_wife in person_id_map:
                    wife_id = person_id_map[original_wife]
                else:
                    try:
                        wife_id = f"I{int(float(original_wife))}"
                    except (ValueError, TypeError):
                        wife_id = ""
            
            # Parse children
            children = []
            if children_str:
                children_list = [c.strip() for c in str(children_str).replace(";", ",").split(",") if c.strip()]
                for child_orig in children_list:
                    if child_orig in person_id_map:
                        children.append(person_id_map[child_orig])
                    else:
                        try:
                            children.append(f"I{int(float(child_orig))}")
                        except (ValueError, TypeError):
                            pass
            
            gen.add_family(ged_family_id, husband_id, wife_id, children, marriage_date, marriage_place)
            print(f"  Added family: {ged_family_id}")
            fam_counter += 1
    else:
        print("[WARNING] 'Families' sheet not found")
    
    # Add header at the beginning
    final_lines = gen.ged_lines
    gen.ged_lines = []
    gen.create_header()
    gen.ged_lines.extend(final_lines)
    
    # Save to file
    gen.save(output_ged_path)
    return True


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python excel_to_ged.py <excel_file> [output_ged_file]")
        print("\nExample:")
        print("  python excel_to_ged.py family_data.xlsx family.ged")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else "family.ged"
    
    # Install openpyxl if needed
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
    
    convert_excel_to_ged(excel_file, output_file)
    print("[OK] Conversion complete!")
