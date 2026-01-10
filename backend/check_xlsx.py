import openpyxl
wb = openpyxl.load_workbook('Family tree.xlsx')
print('Sheet names:', wb.sheetnames)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n{sheet_name} sheet:')
    print(f'  Rows: {ws.max_row}, Columns: {ws.max_column}')
    if ws.max_row > 0:
        headers = [cell.value for cell in ws[1]]
        print('  Headers:', headers)
        if ws.max_row > 1:
            print('  First data row:', [cell.value for cell in ws[2]])
