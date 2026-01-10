"""Create a sample Excel file for testing the GED conversion"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Create workbook
wb = openpyxl.Workbook()

# Remove default sheet and create our sheets
wb.remove(wb.active)

# Create People sheet
ws_people = wb.create_sheet("People")
people_headers = ["ID", "Name", "Sex", "Birth Date", "Birth Place", "Death Date", "Death Place"]
ws_people.append(people_headers)

# Style header
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws_people[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add sample people data
people_data = [
    ["P001", "John Smith", "M", "01/01/1950", "New York, USA", "15/03/2020", "Boston, USA"],
    ["P002", "Mary Johnson", "F", "05/06/1952", "London, UK", "", ""],
    ["P003", "Robert Smith", "M", "20/12/1975", "New York, USA", "", ""],
    ["P004", "Sarah Brown", "F", "10/03/1978", "Chicago, USA", "", ""],
    ["P005", "James Wilson", "M", "25/07/1980", "Los Angeles, USA", "", ""],
    ["P006", "Emily Smith", "F", "14/05/2000", "New York, USA", "", ""],
    ["P007", "Michael Smith", "M", "22/09/2002", "New York, USA", "", ""],
]

for row in people_data:
    ws_people.append(row)

# Adjust column widths for People sheet
for col in ["A", "B", "C", "D", "E", "F", "G"]:
    ws_people.column_dimensions[col].width = 18

# Create Families sheet
ws_families = wb.create_sheet("Families")
families_headers = ["Family ID", "Husband ID", "Wife ID", "Children IDs"]
ws_families.append(families_headers)

# Style header
for cell in ws_families[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add sample family data
families_data = [
    ["F001", "P001", "P002", "P003"],
    ["F002", "P003", "P004", "P006, P007"],
    ["F003", "P005", "", ""],
]

for row in families_data:
    ws_families.append(row)

# Adjust column widths for Families sheet
for col in ["A", "B", "C", "D"]:
    ws_families.column_dimensions[col].width = 20

# Save the workbook
output_path = r"c:\Users\sarun\OneDrive\Documents\family-tree-app\backend\sample_family_data.xlsx"
wb.save(output_path)
print(f"✓ Sample Excel file created: {output_path}")
print("\nSample data:")
print("- People: 7 individuals")
print("- Families: 3 family units")
print("\nYou can use this as a template for your family data!")
